#!/usr/bin/env python3

from pathlib import Path
import argparse
import json
import re
import uuid

from openpyxl import Workbook, load_workbook


AUTHORITY = "CTBC-PROSPECT-BATCH-001"
CONTRACT_VERSION = "1.0"

SOURCE_REQUIRED = {
    "Họ tên",
    "SĐT liên hệ",
    "Mã định danh Bộ GD&ĐT",
}

CANONICAL_HEADERS = [
    "full_name",
    "phone",
    "email",
    "country",
    "program_of_interest",
    "student_number",
    "external_reference",
    "school",
    "province",
    "notes",
    "campaign_source",
]


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value):
    return text(value)


def find_sheet(wb):
    for ws in wb.worksheets:
        for row_number, row in enumerate(
            ws.iter_rows(values_only=True),
            start=1,
        ):
            headers = [normalize_header(v) for v in row]

            if SOURCE_REQUIRED.issubset(set(headers)):
                header_map = {
                    header: index + 1
                    for index, header in enumerate(headers)
                    if header
                }

                return ws, row_number, header_map

    raise ValueError(
        "Unsupported workbook: required source headers not found"
    )


def normalize_phone(value):
    return text(value)


def normalize_student_number(value):
    return text(value)


def output_name(source):
    stem = source.stem

    # Prefer a clean class-oriented filename where the source
    # contains a recognizable class designation.
    cleaned = re.sub(r"[^\wÀ-ỹ]+", "_", stem, flags=re.UNICODE)
    cleaned = cleaned.strip("_")

    return f"CTBC_Prospects_{cleaned}.xlsx"


def convert(source, output_dir):
    wb = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )

    ws, header_row, hm = find_sheet(wb)

    missing = SOURCE_REQUIRED - set(hm)
    if missing:
        raise ValueError(
            "Required headers missing after detection: "
            + ", ".join(sorted(missing))
        )

    rows = []
    duplicate_phones = []
    seen_phones = set()

    # Bound processing to the actual worksheet extent.
    try:
        actual_max_row = ws.calculate_dimension().split(":")[-1]
        actual_max_row = int("".join(ch for ch in actual_max_row if ch.isdigit()))
    except Exception:
        actual_max_row = ws.max_row

    actual_max_row = min(actual_max_row, ws.max_row)

    for row_number in range(header_row + 1, actual_max_row + 1):
        name = text(ws.cell(row_number, hm["Họ tên"]).value)
        phone = normalize_phone(ws.cell(row_number, hm["SĐT liên hệ"]).value)
        student_number = normalize_student_number(
            ws.cell(row_number, hm["Mã định danh Bộ GD&ĐT"]).value
        )

        if not name and not phone and not student_number:
            continue

        if phone:
            if phone in seen_phones:
                duplicate_phones.append(phone)
            seen_phones.add(phone)

        rows.append({
            "full_name": name,
            "phone": phone,
            "email": "",
            "country": "Vietnam",
            "program_of_interest": "",
            "student_number": student_number,
            "external_reference": "",
            "school": "",
            "province": "",
            "notes": "",
            "campaign_source": "",
        })

    output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / output_name(source)

    out = Workbook()
    ws_out = out.active
    ws_out.title = "Prospects"

    ws_out.append(CANONICAL_HEADERS)

    for item in rows:
        ws_out.append([
            item[h] for h in CANONICAL_HEADERS
        ])

    # Preserve phone and student number as TEXT.
    phone_col = CANONICAL_HEADERS.index("phone") + 1
    student_col = CANONICAL_HEADERS.index("student_number") + 1

    for row_number in range(2, ws_out.max_row + 1):
        ws_out.cell(row_number, phone_col).number_format = "@"
        ws_out.cell(row_number, student_col).number_format = "@"

    out.save(output_path)

    return {
        "source_file": source.name,
        "source_sheet": ws.title,
        "header_row": header_row,
        "rows_converted": len(rows),
        "exceptions": [],
        "duplicate_phones": sorted(set(duplicate_phones)),
        "output_file": output_path.name,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert CTBC source prospect workbooks to canonical Prospect List format."
    )

    parser.add_argument(
        "inputs",
        nargs="+",
        type=Path,
    )

    parser.add_argument(
        "--out",
        type=Path,
        default=Path("converted"),
    )

    parser.add_argument(
        "--report",
        type=Path,
        default=None,
    )

    args = parser.parse_args()

    results = []

    for source in args.inputs:
        result = {
            "source_file": source.name,
            "rows_converted": 0,
            "exceptions": [],
            "duplicate_phones": [],
            "output_file": None,
        }

        try:
            converted = convert(source, args.out)
            result.update(converted)
            print(
                f"PASS: {source.name} -> "
                f"{converted['output_file']} "
                f"({converted['rows_converted']} rows)"
            )
        except Exception as exc:
            result["exceptions"] = [str(exc)]
            print(f"FAIL: {source.name}: {exc}")

        results.append(result)

    report = {
        "authority": AUTHORITY,
        "contract_version": CONTRACT_VERSION,
        "canonical_headers": CANONICAL_HEADERS,
        "files": results,
        "total_rows": sum(
            r["rows_converted"] for r in results
        ),
        "failed_files": sum(
            1 for r in results if r["exceptions"]
        ),
    }

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(
            json.dumps(
                report,
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    print()
    print("AUTHORITY=" + AUTHORITY)
    print("INPUT_FILES=" + str(len(results)))
    print("FAILED_FILES=" + str(report["failed_files"]))
    print("TOTAL_ROWS=" + str(report["total_rows"]))

    if report["failed_files"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
